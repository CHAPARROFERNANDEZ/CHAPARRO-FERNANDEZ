import os
import calendar
from datetime import datetime
import pandas as pd

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


ARCHIVO = "inversiones.xlsx"
HOJA = "INVERSIONES"
CARPETA_EXTRACTOS = "extractos"


# =========================
# PREGUNTAS DINÁMICAS
# =========================
print("\n=== GENERADOR DE EXTRACTOS ===")

modo = input("¿Quieres generar extracto de TODOS o de UN inversor concreto? (todos/uno): ").strip().lower()

inversor_elegido = None

if modo == "uno":
    inversor_elegido = input("Escribe el nombre del inversor exactamente como sale en Excel: ").strip()

año = int(input("¿De qué año quieres el extracto? Ejemplo 2026: ").strip())
mes = int(input("¿Hasta qué mes quieres el extracto? Ejemplo 3 para marzo: ").strip())

ultimo_dia = calendar.monthrange(año, mes)[1]
FECHA_CORTE = datetime(año, mes, ultimo_dia)

print(f"\nFecha de corte seleccionada: {FECHA_CORTE.strftime('%d/%m/%Y')}")


def convertir_fecha(valor):
    if pd.isna(valor) or valor == "":
        return pd.NaT
    return pd.to_datetime(valor, dayfirst=True, errors="coerce")


def generar_meses(fecha_inicio, fecha_fin):
    meses = []
    actual = datetime(fecha_inicio.year, fecha_inicio.month, 1)
    fin_mes = datetime(fecha_fin.year, fecha_fin.month, 1)

    while actual <= fin_mes:
        meses.append(actual)

        if actual.month == 12:
            actual = datetime(actual.year + 1, 1, 1)
        else:
            actual = datetime(actual.year, actual.month + 1, 1)

    return meses


def formatear_excel(nombre_archivo, inversor, fecha_corte):
    wb = load_workbook(nombre_archivo)

    azul = "1F4E78"
    azul_claro = "D9EAF7"
    verde = "E2F0D9"
    blanco = "FFFFFF"

    borde_fino = Side(style="thin", color="D9D9D9")
    borde = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(vertical="center")
                cell.border = borde

        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 22

    ws = wb["RESUMEN"]
    ws.insert_rows(1, 5)

    ws["A1"] = "EXTRACTO DE INVERSIÓN"
    ws["A1"].font = Font(name="Calibri", size=20, bold=True, color=blanco)
    ws["A1"].fill = PatternFill("solid", fgColor=azul)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:D1")

    ws["A3"] = "Inversor"
    ws["B3"] = inversor
    ws["A4"] = "Fecha de corte"
    ws["B4"] = fecha_corte.strftime("%d/%m/%Y")

    for cell in ["A3", "A4"]:
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill("solid", fgColor=azul_claro)

    for row in range(6, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).alignment = Alignment(horizontal="center")

    for cell in ws[6]:
        cell.font = Font(bold=True, color=blanco)
        cell.fill = PatternFill("solid", fgColor=azul)
        cell.alignment = Alignment(horizontal="center")

    for row in range(7, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=verde)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 26

    ws = wb["TOTALES_MES"]
    ws.insert_rows(1, 3)

    ws["A1"] = "RESUMEN MENSUAL"
    ws["A1"].font = Font(size=18, bold=True, color=blanco)
    ws["A1"].fill = PatternFill("solid", fgColor=azul)
    ws.merge_cells("A1:B1")
    ws["A1"].alignment = Alignment(horizontal="center")

    for cell in ws[4]:
        cell.font = Font(bold=True, color=blanco)
        cell.fill = PatternFill("solid", fgColor=azul)
        cell.alignment = Alignment(horizontal="center")

    for row in range(5, ws.max_row + 1):
        ws.cell(row, 2).number_format = '#,##0.00 €'

        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18

    ws = wb["DETALLE"]
    ws.insert_rows(1, 3)

    ws["A1"] = "DETALLE DEL EXTRACTO"
    ws["A1"].font = Font(size=18, bold=True, color=blanco)
    ws["A1"].fill = PatternFill("solid", fgColor=azul)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    ws["A1"].alignment = Alignment(horizontal="center")

    for cell in ws[4]:
        cell.font = Font(bold=True, color=blanco)
        cell.fill = PatternFill("solid", fgColor=azul)
        cell.alignment = Alignment(horizontal="center")

    for row in range(5, ws.max_row + 1):
        fill = azul_claro if row % 2 == 0 else blanco

        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=fill)
            ws.cell(row, col).alignment = Alignment(horizontal="center")

        if ws.max_column >= 8:
            ws.cell(row, 8).number_format = '#,##0.00 €'

        if ws.max_column >= 11:
            ws.cell(row, 11).number_format = '#,##0.00 €'

    anchos = {
        "A": 24,
        "B": 16,
        "C": 18,
        "D": 18,
        "E": 20,
        "F": 14,
        "G": 18,
        "H": 18,
        "I": 16,
        "J": 12,
        "K": 18,
    }

    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    wb.save(nombre_archivo)


# =========================
# CARGAR EXCEL
# =========================
df = pd.read_excel(ARCHIVO, sheet_name=HOJA)
df.columns = df.columns.str.strip()


# =========================
# NORMALIZAR COLUMNAS
# =========================
columnas_texto = [
    "inversor",
    "tipo_inversion",
    "subtipo_inversion",
    "nombre_activo",
    "tipo_operacion",
    "capital_nuevo_real",
]

for col in columnas_texto:
    df[col] = df[col].fillna("").astype(str).str.strip()

df["tipo_inversion"] = df["tipo_inversion"].str.upper()
df["subtipo_inversion"] = df["subtipo_inversion"].str.upper()
df["nombre_activo"] = df["nombre_activo"].str.upper()
df["tipo_operacion"] = df["tipo_operacion"].str.upper()
df["capital_nuevo_real"] = df["capital_nuevo_real"].str.upper()

df["fecha_inversion"] = df["fecha_inversion"].apply(convertir_fecha)
df["fecha_final_inversion"] = df["fecha_final_inversion"].apply(convertir_fecha)

df["capital_invertido"] = pd.to_numeric(df["capital_invertido"], errors="coerce").fillna(0)
df["interes_inversor_anual"] = pd.to_numeric(df["interes_inversor_anual"], errors="coerce").fillna(0)


# =========================
# FILTRAR CAPITAL REAL
# =========================
df = df[df["capital_nuevo_real"].isin(["SI", ""])].copy()


# =========================
# FILTRAR INVERSOR SI PROCEDE
# =========================
if modo == "uno":
    df = df[df["inversor"].str.upper() == inversor_elegido.upper()].copy()

    if df.empty:
        print(f"No se ha encontrado ningún inversor llamado: {inversor_elegido}")
        exit()


# =========================
# CALCULAR EXTRACTOS
# =========================
filas = []

for _, row in df.iterrows():
    inversor = row["inversor"]
    id_inversion = row["id_inversion"]
    tipo_inversion = row["tipo_inversion"]
    subtipo_inversion = row["subtipo_inversion"]
    nombre_activo = row["nombre_activo"]

    fecha_inicio = row["fecha_inversion"]
    fecha_final_excel = row["fecha_final_inversion"]

    capital = row["capital_invertido"]
    interes_anual = row["interes_inversor_anual"]
    interes_mensual = capital * interes_anual / 12

    if pd.isna(fecha_inicio):
        continue

    if subtipo_inversion == "ESTRUCTURADA":
        fecha_fin = FECHA_CORTE
    else:
        if pd.isna(fecha_final_excel):
            fecha_fin = FECHA_CORTE
        else:
            fecha_fin = min(fecha_final_excel.to_pydatetime(), FECHA_CORTE)

    if fecha_inicio > fecha_fin:
        continue

    fecha_inicio_mes = datetime(fecha_inicio.year, fecha_inicio.month, 1)
    meses = generar_meses(fecha_inicio_mes, fecha_fin)

    for mes_inicio in meses:
        año_mes = mes_inicio.year
        mes_num = mes_inicio.month
        dias_mes = calendar.monthrange(año_mes, mes_num)[1]

        inicio_mes = datetime(año_mes, mes_num, 1)
        fin_mes = datetime(año_mes, mes_num, dias_mes)

        inicio_calculo = max(fecha_inicio.to_pydatetime(), inicio_mes)
        fin_calculo = min(fecha_fin, fin_mes)

        if inicio_calculo <= fin_calculo:
            dias_devengados = (fin_calculo - inicio_calculo).days + 1
            interes_mes = round(interes_mensual * dias_devengados / dias_mes, 2)

            filas.append({
                "inversor": inversor,
                "id_inversion": id_inversion,
                "tipo_inversion": tipo_inversion,
                "subtipo_inversion": subtipo_inversion,
                "nombre_activo": nombre_activo,
                "mes": f"{mes_num:02d}/{año_mes}",
                "fecha_inversion": fecha_inicio.strftime("%d/%m/%Y"),
                "capital_invertido": capital,
                "dias_devengados": dias_devengados,
                "dias_mes": dias_mes,
                "interes_mes": interes_mes,
            })


# =========================
# EXPORTAR RESULTADO
# =========================
resultado = pd.DataFrame(filas)

if resultado.empty:
    print("No se han generado filas. Revisa el Excel o la fecha seleccionada.")
else:
    resultado["mes_orden"] = pd.to_datetime("01/" + resultado["mes"], dayfirst=True)

    resultado = resultado.sort_values(
        by=["inversor", "mes_orden", "id_inversion"]
    ).drop(columns=["mes_orden"]).reset_index(drop=True)

    os.makedirs(CARPETA_EXTRACTOS, exist_ok=True)

    for inversor, grupo in resultado.groupby("inversor"):
        detalle = grupo.copy()

        totales_mes = (
            detalle.groupby("mes", as_index=False)["interes_mes"]
            .sum()
            .rename(columns={"interes_mes": "total_mes"})
        )

        totales_mes["mes_orden"] = pd.to_datetime("01/" + totales_mes["mes"], dayfirst=True)

        totales_mes = totales_mes.sort_values(
            by="mes_orden"
        ).drop(columns=["mes_orden"]).reset_index(drop=True)

        capital_total = (
            detalle.groupby("id_inversion")["capital_invertido"]
            .first()
            .sum()
        )

        resumen = pd.DataFrame([{
            "inversor": inversor,
            "fecha_corte": FECHA_CORTE.strftime("%d/%m/%Y"),
            "capital_total": round(capital_total, 2),
            "total_intereses_acumulados": round(detalle["interes_mes"].sum(), 2),
        }])

        nombre_archivo = os.path.join(
            CARPETA_EXTRACTOS,
            f"extracto_{inversor.upper().replace(' ', '_')}_{FECHA_CORTE.strftime('%d%m%Y')}.xlsx"
        )

        with pd.ExcelWriter(nombre_archivo, engine="openpyxl") as writer:
            resumen.to_excel(writer, sheet_name="RESUMEN", index=False)
            totales_mes.to_excel(writer, sheet_name="TOTALES_MES", index=False)
            detalle.to_excel(writer, sheet_name="DETALLE", index=False)

        formatear_excel(nombre_archivo, inversor, FECHA_CORTE)

        print(f"✅ Extracto generado: {nombre_archivo}")

    print("✅ Todos los extractos generados correctamente.")